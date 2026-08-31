"""A table of rows, from whichever file the operator happens to have.

The reference arrives as a spreadsheet, the pipeline writes a CSV, and the two
have to be compared to each other. Reading both here means the comparison never
has to know which one it was handed, and the operator never has to export a
sheet to CSV before asking a question about it.

Rows are keyed by document number rather than by the filename cell, because the
two files spell that cell differently — ``100001.pdf`` in the spreadsheet,
``100001`` in ours — and a key that disagrees is a comparison of nothing. The
number is the identity; everything around it is formatting.
"""

from __future__ import annotations

import csv
import re
from pathlib import Path

csv.field_size_limit(10**8)

#: The document number inside whatever the filename cell holds.
_NUMBER = re.compile(r"^\d{5,7}(?:\.\d+)?")

#: The column both files carry the filename in, trailing space and all.
FILENAME = "ชื่อไฟล์ "


def document_of(value: object) -> str:
    """``100001`` from ``100001.pdf``, ``100001.0``, or ``100001``."""
    found = _NUMBER.search(str(value or ""))
    return found.group(0) if found else ""


def read(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    """(header, rows) from a ``.csv`` or ``.xlsx``, every cell a string."""
    if path.suffix.lower() in (".xlsx", ".xlsm"):
        return _from_workbook(path)
    return _from_csv(path)


def _from_csv(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        table = list(csv.reader(handle))
    if not table:
        return [], []
    header = table[0]
    return header, [
        {name: (row[at] if at < len(row) else "") for at, name in enumerate(header)}
        for row in table[1:]
    ]


def _from_workbook(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover - depends on the environment
        raise SystemExit(
            "ต้องติดตั้ง openpyxl ก่อนจึงจะอ่าน .xlsx ได้:  pip install openpyxl"
        ) from exc

    book = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows = book[book.sheetnames[0]].iter_rows(values_only=True)
    header = [("" if cell is None else str(cell)) for cell in next(rows, ())]
    table = [
        {name: ("" if value is None else str(value))
         for name, value in zip(header, row, strict=False)}
        for row in rows
    ]
    book.close()
    return header, table


def by_document(path: Path) -> tuple[list[str], dict[str, dict[str, str]]]:
    """Rows indexed by document number, first row wins on a duplicate."""
    header, rows = read(path)
    indexed: dict[str, dict[str, str]] = {}
    for row in rows:
        number = document_of(row.get(FILENAME) or next(iter(row.values()), ""))
        if number:
            indexed.setdefault(number, row)
    return header, indexed
