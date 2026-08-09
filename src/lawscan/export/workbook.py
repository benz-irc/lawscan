"""The comparison as a spreadsheet, with the arithmetic shown.

A score is only worth as much as the reader's ability to disagree with it. This
writes two sheets and they answer different questions: the first says how far
apart the two files are, and the second says which cells make up that distance
and why each one was counted the way it was.

The ``เหตุผลที่นับแบบนี้`` column is the point of the whole file. The comparison
forgives differences Thai does not use to carry meaning — spacing, Thai against
Arabic digits, a vowel the extraction broke, the order of a list — and a score
that quietly forgives things is a score nobody can audit. Every cell counted as
agreeing names the difference it was forgiven, so a reader who thinks a fold
went too far can find every cell it touched and say so.
"""

from __future__ import annotations

import json
from dataclasses import dataclass
from pathlib import Path

from lawscan import adjudicate, sheet, verdict
from lawscan.diff import (
    EXTERNAL, LISTS, LITERAL, PROSE, UNSCORED, compare_cell, match_reason,
)
from lawscan.export.columns import COLUMNS

#: How each verdict is written for a reader. ``ใกล้เคียง`` rather than "partial"
#: because the sheet is read by the person who has to decide whether to act on
#: it, and "close" is the word that tells them there is a cell here worth
#: looking at rather than a cell to rewrite.
VERDICTS: dict[str, str] = {
    "exact": "ตรง",
    "partial": "ใกล้เคียง",
    "wrong": "ไม่ตรง",
    "blank": "ว่างทั้งคู่",
}

#: Sheet titles. Excel caps these at 31 characters and silently renames longer
#: ones, which would break a formula pointing at them.
SUMMARY = "สรุป"
DETAIL = "รายช่อง"

_HEADERS = (
    "เอกสาร", "คอลัมน์", "ผล", "เหตุผลที่นับแบบนี้", "ชนิดช่อง",
    "ของผู้ดูแล", "ของเรา", "ฝั่งไหนดีกว่า", "ทำไมถึงตัดสินแบบนั้น", "ที่มาของเรา",
)

#: Column widths, in the order above. Chosen so the two value columns are the
#: wide ones: they are what the reader is here to compare.
_WIDTHS = (10, 34, 11, 26, 12, 60, 60, 30, 60, 20)

#: (fill, text) per verdict — Excel's own Good/Neutral/Bad palette. Borrowed
#: rather than invented because the reader has seen these three colours mean
#: exactly this in every other spreadsheet they have opened, and a legend
#: nobody has to read is worth more than a nicer green.
#:
#: The fill goes on the verdict cell and, paler, across the row. Colour on the
#: verdict alone leaves the eye hunting for which row it belonged to once the
#: sheet is filtered; colour across the row at full strength makes the text
#: hard to read. Two strengths of the same hue is what makes a long sheet
#: scannable without turning it into a poster.
BANDS: dict[str, tuple[str, str, str]] = {
    "exact":   ("C6EFCE", "006100", "EFF7EE"),
    "partial": ("FFEB9C", "9C5700", "FFF9E8"),
    "wrong":   ("FFC7CE", "9C0006", "FDF0F1"),
    "blank":   ("EDEDED", "7F7F7F", "FAFAFA"),
}

#: Where a column's score stops being good news. The two thresholds are the
#: same ones a reader uses without being told: near enough to leave alone, and
#: bad enough to be this week's work.
_GOOD, _FAIR = 0.80, 0.50

_HEADER_FILL = "1F3864"
_SECTION_FILL = "D9E2F3"


def kind_of(column: str) -> str:
    """What sort of cell this is, so a reader knows what the verdict means."""
    if column in PROSE:
        return "ข้อความยาว"
    if column in EXTERNAL:
        return "นอกเอกสาร"
    if column in LISTS:
        return "รายการ"
    return "ค่าเดียว"


@dataclass(frozen=True, slots=True)
class Cell:
    document: str
    column: str
    theirs: str
    ours: str
    verdict: str
    reason: str
    origin: str

    @property
    def counted(self) -> bool:
        return self.column not in UNSCORED and self.verdict != "blank"


def _origins(workdir: Path | None) -> dict[str, dict[str, str]]:
    """Which half of the pipeline produced each of our cells, per document.

    Read from the run's own folders rather than recomputed. A cell that a rule
    filled and a cell the model filled fail for different reasons and are fixed
    in different files, and that is the first thing anyone asks about a miss.
    """
    if workdir is None or not workdir.is_dir():
        return {}
    found: dict[str, dict[str, str]] = {}
    for folder in sorted(workdir.iterdir()):
        record = folder / "row.json"
        if not record.is_file():
            continue
        try:
            found[folder.name] = json.loads(record.read_text(encoding="utf-8")).get(
                "sources", {}
            )
        except ValueError:
            continue
    return found


_ORIGIN_WORDS = {"rule": "กฎ (regex/ทะเบียน)", "llm": "โมเดล"}


def _origin_text(origin: str) -> str:
    if not origin:
        return "—"
    head, _, tail = origin.partition(":")
    word = _ORIGIN_WORDS.get(head, head)
    return f"{word} · {tail}" if tail else word


def cells(expected: Path, ours: Path, *, workdir: Path | None = None) -> list[Cell]:
    """Every shared cell of every shared document, in the export's order."""
    their_header, theirs = sheet.by_document(expected)
    our_header, mine = sheet.by_document(ours)
    shared = sorted(set(theirs) & set(mine))
    columns = [c for c in COLUMNS if c in their_header and c in our_header]
    origins = _origins(workdir)

    found: list[Cell] = []
    for document in shared:
        source = origins.get(document, {})
        for column in columns:
            left = (theirs[document].get(column) or "").strip()
            right = (mine[document].get(column) or "").strip()
            outcome = compare_cell(column, left, right)
            found.append(Cell(
                document=document,
                column=column,
                theirs=left,
                ours=right,
                verdict=outcome,
                reason=(match_reason(left, right, column) if outcome == "exact"
                        else _near_reason(column, left, right, outcome)),
                origin=source.get(column, ""),
            ))
    return found


def _near_reason(column: str, theirs: str, ours: str, outcome: str) -> str:
    """Why a cell was not counted as agreeing, in one phrase."""
    if outcome == "blank":
        return "ทั้งคู่ไม่มีคำตอบ"
    if not theirs:
        return "เฉลยเว้นว่าง เรากรอก"
    if not ours:
        return "เฉลยมีคำตอบ เราเว้นว่าง"
    if outcome == "partial":
        return "ซ้อนกันบางส่วน" if column in LISTS else "ฝั่งหนึ่งคลุมอีกฝั่ง"
    return "คนละคำตอบ"


@dataclass(frozen=True, slots=True)
class Tally:
    exact: int = 0
    partial: int = 0
    wrong: int = 0
    blank: int = 0

    @property
    def scored(self) -> int:
        return self.exact + self.partial + self.wrong

    @property
    def credit(self) -> float:
        return self.exact + 0.5 * self.partial

    def plus(self, outcome: str) -> Tally:
        counts = {name: getattr(self, name)
                  for name in ("exact", "partial", "wrong", "blank")}
        counts[outcome] += 1
        return Tally(**counts)


def _tally(found: list[Cell]) -> Tally:
    total = Tally()
    for cell in found:
        if cell.column in UNSCORED:
            continue
        total = total.plus(cell.verdict)
    return total


def _share(part: float, whole: float) -> str:
    return f"{part / whole:.1%}" if whole else "—"


@dataclass(frozen=True, slots=True)
class Line:
    """One row of the summary, and how it should be painted.

    The style travels with the row rather than being worked out afterwards
    from what the row says. A painter that has to recognise ``"ไม่ตรง"`` in
    column A to decide on red is a painter that silently stops working the day
    somebody rewords a label.
    """

    values: list[object]
    style: str = ""


def per_column(found: list[Cell]) -> dict[str, Tally]:
    """Each column's own four numbers, worst column first when sorted."""
    tallies: dict[str, Tally] = {}
    for cell in found:
        tallies[cell.column] = tallies.get(cell.column, Tally()).plus(cell.verdict)
    return tallies


def _summary_rows(found: list[Cell]) -> list[Line]:
    """The headline, the folds that were applied, and the per-column table."""
    total = _tally(found)
    documents = len({c.document for c in found})

    lines = [
        Line(["สรุปผลการเทียบ"], "title"),
        Line(["เอกสารที่เทียบได้", documents]),
        Line(["ช่องที่นับคะแนน", total.scored]),
        Line([]),
        Line(["ผล", "จำนวนช่อง", "คิดเป็น"], "head"),
        Line(["ตรง", total.exact, _share(total.exact, total.scored)], "exact"),
        Line(["ใกล้เคียง", total.partial, _share(total.partial, total.scored)], "partial"),
        Line(["ไม่ตรง", total.wrong, _share(total.wrong, total.scored)], "wrong"),
        Line(["ว่างทั้งคู่ (ไม่นับ)", total.blank, "—"], "blank"),
        Line(["คะแนนรวม (ใกล้เคียงนับครึ่ง)", total.credit,
              _share(total.credit, total.scored)], "total"),
        Line([]),
        Line(["ช่องที่นับว่า 'ตรง' ตรงเพราะอะไร", "จำนวนช่อง"], "head"),
    ]

    reasons: dict[str, int] = {}
    for cell in found:
        if cell.verdict == "exact" and cell.column not in UNSCORED:
            reasons[cell.reason] = reasons.get(cell.reason, 0) + 1
    for reason, count in sorted(reasons.items(), key=lambda kv: -kv[1]):
        # The literal matches need no defending; everything below them is a
        # difference this file decided to forgive, and is tinted to say so.
        literal = reason == LITERAL
        lines.append(Line([reason, count], "" if literal else "forgiven"))

    lines += [
        Line([]),
        Line(["คอลัมน์", "ชนิด", "ตรง", "ใกล้เคียง", "ไม่ตรง", "ว่างทั้งคู่",
              "คะแนน", "นับไหม"], "head"),
    ]
    # Worst first, but the uncounted columns after all the counted ones. The
    # table is a work queue: ``ลิงค์PDF`` disagrees on all 240 rows because
    # nothing in a document carries the Gazette's own URL, and sorting that to
    # the top puts an unfixable row where the next job should be.
    for column, tally in sorted(
        per_column(found).items(),
        key=lambda kv: (kv[0] in UNSCORED, -(kv[1].wrong + 0.5 * kv[1].partial), kv[0]),
    ):
        counted = column not in UNSCORED
        share = tally.credit / tally.scored if tally.scored else None
        lines.append(Line(
            [
                column.strip(), kind_of(column),
                tally.exact, tally.partial, tally.wrong, tally.blank,
                _share(tally.credit, tally.scored),
                "นับ" if counted else "ไม่นับ",
            ],
            "" if not counted or share is None
            else "exact" if share >= _GOOD
            else "partial" if share >= _FAIR
            else "wrong",
        ))
    return lines


def _text_of(workdir: Path | None, document: str) -> str:
    """The document's own words, when the run kept them."""
    if workdir is None:
        return ""
    saved = workdir / document / "text.txt"
    return saved.read_text(encoding="utf-8") if saved.is_file() else ""


def judge(cell: Cell, text: str) -> tuple[str, str]:
    """(which side is better, why) for a cell the two files disagree about.

    Three different questions, and which one applies depends on the column.
    The Gazette line is printed on the page, so the document settles it. A
    summary is nobody's to settle by string search, but which side's words are
    in the document at all is worth saying and is not the same claim. Everything
    else is decided on the shape of the two cells, which claims least of all.
    """
    if cell.verdict == "exact":
        return "", ""
    if cell.column == "ข้อมูลแหล่งที่มา" and text:
        side, reason = adjudicate.source_line(cell.theirs, cell.ours, text)
        if reason:
            return side, reason
    if cell.column in PROSE and text and cell.theirs and cell.ours:
        return adjudicate.grounding(cell.theirs, cell.ours, text)
    return (
        verdict.better(cell.column, cell.theirs, cell.ours),
        verdict.why(cell.column, cell.theirs, cell.ours, origin=cell.origin),
    )


def write(expected: Path, ours: Path, out: Path, *,
          workdir: Path | None = None) -> Tally:
    """Both sheets to ``out``. Returns the headline tally for printing."""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover - depends on the environment
        raise SystemExit(
            "ต้องติดตั้ง openpyxl ก่อนจึงจะเขียน .xlsx ได้:  pip install openpyxl"
        ) from exc

    found = cells(expected, ours, workdir=workdir)
    book = openpyxl.Workbook()

    def solid(colour: str) -> PatternFill:
        return PatternFill("solid", fgColor=colour)

    page = book.active
    page.title = SUMMARY
    for line in _summary_rows(found):
        page.append(line.values)
        if not line.values:
            continue
        row = page[page.max_row]
        if line.style == "title":
            row[0].font = Font(bold=True, size=14, color=_HEADER_FILL)
        elif line.style == "head":
            for column in row:
                column.font = Font(bold=True, color="FFFFFF")
                column.fill = solid(_HEADER_FILL)
        elif line.style == "total":
            for column in row:
                column.font = Font(bold=True)
                column.fill = solid(_SECTION_FILL)
        elif line.style == "forgiven":
            # Not a wrong answer and not a plain match either: a difference
            # this file chose to overlook. Tinted so the reader can see at a
            # glance how much of the score rests on that choice.
            for column in row:
                column.fill = solid("EAF1F8")
                column.font = Font(color="1F3864")
        elif line.style in BANDS:
            strong, ink, pale = BANDS[line.style]
            row[0].fill = solid(strong)
            row[0].font = Font(bold=True, color=ink)
            for column in row[1:]:
                column.fill = solid(pale)
    for width, letter in zip((44, 14, 12, 12, 12, 14, 12, 10),
                             "ABCDEFGH", strict=False):
        page.column_dimensions[letter].width = width
    page.freeze_panes = "A2"

    detail = book.create_sheet(DETAIL)
    detail.append(list(_HEADERS))

    # Built once and shared. A fresh Font per cell on eight thousand rows is
    # eight thousand style records in the file, which openpyxl will write and
    # Excel will then be slow opening.
    verdict_style = {
        name: (solid(strong), Font(bold=True, color=ink), solid(pale))
        for name, (strong, ink, pale) in BANDS.items()
    }
    faint = Font(color="909090")
    wrap = Alignment(wrap_text=True, vertical="top")
    at = {name: position for position, name in enumerate(_HEADERS)}

    texts: dict[str, str] = {}
    for cell in found:
        if cell.document not in texts:
            texts[cell.document] = _text_of(workdir, cell.document)
        side, reason = judge(cell, texts[cell.document])
        detail.append([
            cell.document, cell.column.strip(), VERDICTS[cell.verdict], cell.reason,
            kind_of(cell.column), cell.theirs, cell.ours, side, reason,
            _origin_text(cell.origin),
        ])
        row = detail[detail.max_row]
        strong, ink, pale = verdict_style[cell.verdict]
        mark = row[at["ผล"]]
        mark.fill, mark.font = strong, ink
        for position in (at["ของผู้ดูแล"], at["ของเรา"]):
            row[position].fill = pale
            row[position].alignment = wrap
        if cell.column in UNSCORED:
            # Shown because the reader wants to see it, not counted because
            # nothing in the document can settle it. Greyed so the two are not
            # confused when the sheet is being skimmed.
            row[at["คอลัมน์"]].font = faint

    header = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor=_HEADER_FILL)
    for column in detail[1]:
        column.font = header
        column.fill = fill
    for width, position in zip(_WIDTHS, range(1, len(_HEADERS) + 1), strict=False):
        detail.column_dimensions[get_column_letter(position)].width = width
    # Frozen past the document and column, so a row keeps saying which cell it
    # is about while the reader scrolls right into the two values.
    detail.freeze_panes = "C2"
    detail.auto_filter.ref = detail.dimensions

    out.parent.mkdir(parents=True, exist_ok=True)
    book.save(out)
    return _tally(found)


__all__ = [
    "BANDS", "Cell", "Line", "Tally", "cells", "judge", "kind_of",
    "per_column", "write", "VERDICTS",
]
